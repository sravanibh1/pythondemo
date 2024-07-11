import time
import openpyxl
from selenium import webdriver
from sampleproject.autoproject.pages.basepage import BasePage
from sampleproject.autoproject.pages.detailspage import DetailsPage


def driver_setup():
    driver = webdriver.Chrome()
    driver.maximize_window()
    bp = BasePage(driver)
    bp.open_url("https://testautomationpractice.blogspot.com/")
    return driver


def excel_data():
    workbook = openpyxl.load_workbook("C:\\Users\\Sravani Bhogadi\\Desktop\\Book1.xlsx")
    sheet = workbook.active
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        email = sheet.cell(row=row, column=2).value
        phone = sheet.cell(row=row, column=3).value
        address = sheet.cell(row=row, column=4).value
        gender = sheet.cell(row=row, column=5).value
        country = sheet.cell(row=row, column=6).value
        day = sheet.cell(row=row, column=7).value
        color = sheet.cell(row=row, column=8).value
        bookname = sheet.cell(row=row, column=9).value
        subject = sheet.cell(row=row, column=10).value
        product = sheet.cell(row=row, column=11).value
        search_value = sheet.cell(row=row, column=12).value
        filling(name, email, phone, address, gender, country, day, color, bookname, subject, product, search_value)


def filling(name, email, phone, address, gender, country, day, color, bookname, subject, product, search):
    driver = driver_setup()
    dp = DetailsPage(driver)
    dp.enter_name(name)
    dp.enter_email(email)
    dp.enter_phone(phone)
    dp.enter_address(address)
    dp.select_gender(gender)
    dp.enter_country(country)
    dp.select_day(day)
    dp.enter_colour(color)
    dp.enter_date('October', '2024', '2')
    dp.get_price(bookname)
    dp.get_author(subject)
    dp.click_product(product)
    dp.search(search)
    dp.alerts()
    dp.double()
    dp.drag_drop()
    dp.slider(50, 0)
    driver.quit()

# filling('male',  'monday')
# filling('female', 'sunday')


excel_data()

